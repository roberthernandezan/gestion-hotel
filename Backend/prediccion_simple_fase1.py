import os
import django
import pandas as pd
from django.db.models import Sum, F, Min, Max
import warnings
from itertools import product  
from openpyxl.utils import get_column_letter  

warnings.filterwarnings("ignore")  

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "hotel_management.settings")
django.setup()

from gestion.models import OrdenElementos, CocktailIngredientes, Ingredientes

def generar_consumos_historicos_excel():
    # Obtener fecha de inicio (primer registro) y de finalización (último registro)
    fecha_inicio = OrdenElementos.objects.aggregate(primera_fecha=Min("fechaorden"))["primera_fecha"]
    fecha_fin = OrdenElementos.objects.aggregate(ultima_fecha=Max("fechaorden"))["ultima_fecha"]

    if not fecha_inicio or not fecha_fin:
        print("No hay registros en la base de datos para calcular consumos.")
        return

    print(f"Fecha de inicio detectada automáticamente: {fecha_inicio}")
    print(f"Fecha de finalización detectada automáticamente: {fecha_fin}")

    consumo_total = {}

    # Cálculo de consumo para registros directos (no cocktail)
    datos_directos = (
        OrdenElementos.objects
        .filter(escocktail=False, fechaorden__range=[fecha_inicio, fecha_fin])
        .annotate(
            año=F('fechaorden__year'),
            mes=F('fechaorden__month'),
            ingrediente_nombre=F('id_ingredientes__nombre')
        )
        .values('ingrediente_nombre', 'año', 'mes')
        .annotate(total_litros=Sum(F('cantidad') * F('id_ingredientes__litrosporunidad')))
    )

    for dato in datos_directos:
        key = (dato['ingrediente_nombre'], dato['año'], dato['mes'])
        consumo_total[key] = consumo_total.get(key, 0) + dato['total_litros']

    # Cálculo de consumo para registros de cocktails
    datos_cocktails = OrdenElementos.objects.filter(
        escocktail=True,
        fechaorden__range=[fecha_inicio, fecha_fin]
    )

    for cocktail in datos_cocktails:
        receta = CocktailIngredientes.objects.filter(
            id_cocktail=cocktail.id_cocktail,
            activo=True
        )
        for ingrediente in receta:
            key = (ingrediente.id_ingredientes.nombre, cocktail.fechaorden.year, cocktail.fechaorden.month)
            cantidad_usada = cocktail.cantidad * ingrediente.cantidad
            consumo_total[key] = consumo_total.get(key, 0) + cantidad_usada

    # Obtener todos los ingredientes existentes en la base de datos
    ingredientes = Ingredientes.objects.values_list("nombre", flat=True).distinct()
    fecha_fin_datetime = pd.to_datetime(fecha_fin)
    años = range(fecha_inicio.year, fecha_fin_datetime.year + 1)
    meses = range(1, 13)

    combinaciones = []
    for año in años:
        if año == fecha_fin_datetime.year:
            meses_hasta_fin = range(1, fecha_fin_datetime.month + 1)
            combinaciones.extend(product(ingredientes, [año], meses_hasta_fin))
        else:
            combinaciones.extend(product(ingredientes, [año], meses))

    datos_consumo = [
        {
            "Ingrediente": ingrediente,
            "Año": año,
            "Mes": mes,
            "Consumo Total": consumo_total.get((ingrediente, año, mes), 0)
        }
        for ingrediente, año, mes in combinaciones
    ]

    df = pd.DataFrame(datos_consumo)
    df.sort_values(by=['Ingrediente', 'Año', 'Mes'], inplace=True)

    # Creamos la tabla pivot: filas = (Ingrediente, Año) y columnas = Mes (1 a 12)
    pivot_table = df.pivot(index=["Ingrediente", "Año"], columns="Mes", values="Consumo Total")
    pivot_table = pivot_table.reindex(columns=range(1, 13), fill_value=0)
    pivot_table = pivot_table.astype(float)

    # Renombrar columnas a abreviaturas de 3 letras
    meses_abreviados = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
        5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
    }
    pivot_table.rename(columns=meses_abreviados, inplace=True)

    # --- Bloque para generar e imprimir vectores de consumos históricos ---
    periodos = []
    current_year = fecha_inicio.year
    current_month = fecha_inicio.month
    while current_year < fecha_fin.year or (current_year == fecha_fin.year and current_month <= fecha_fin.month):
        periodos.append((current_year, current_month))
        if current_month == 12:
            current_year += 1
            current_month = 1
        else:
            current_month += 1

    vectores_consumo = {}
    for ing in ingredientes:
        vector_consumo = [float(consumo_total.get((ing, año, mes), 0)) for (año, mes) in periodos]
        vectores_consumo[ing] = vector_consumo
        print(f"{ing}: {vector_consumo}")    

    # Exportar la tabla pivot a Excel con autoajuste de columnas
    ruta_archivo = "consumo_historico.xlsx"
    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        pivot_table.to_excel(writer, sheet_name="Consumos", index=True)
        worksheet = writer.sheets["Consumos"]
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    print(f"Archivo generado: {ruta_archivo}")
    return fecha_inicio, fecha_fin, vectores_consumo

# Ejecutar la función histórica y obtener vectores
fecha_inicio, fecha_fin, vectores_consumo = generar_consumos_historicos_excel()

# ========================================================================================

import numpy as np
from statsmodels.tsa.statespace.sarimax import SARIMAX

n_pronosticos = 12
predicciones_list = []
meses_abreviados = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
}

for ing, vector in vectores_consumo.items():
    # Crear índice de fechas para la serie histórica
    index_historico = pd.date_range(start=fecha_inicio, periods=len(vector), freq='MS')
    serie = pd.Series(data=vector, index=index_historico)
    
    print(f"\n--- Serie histórica para {ing} ---")
    print(serie.head())
    print("...")
    print(serie.tail())
    ultimo_original = serie.index[-1]
    
    # Transformación logarítmica para estabilizar la varianza
    serie_log = np.log1p(serie)
    
    # Ajustar modelo SARIMAX sobre la serie
    try:
        modelo = SARIMAX(
            serie_log,
            order=(1, 1, 1),
            seasonal_order=(1, 1, 1, 12),
            enforce_stationarity=False,
            enforce_invertibility=False
        )
        resultado = modelo.fit(disp=False)
        print(f"Modelo ajustado correctamente para {ing}.")
    except Exception as e:
        print(f"Error al ajustar el modelo para {ing}: {e}")
        continue
    
    # Definir el inicio del pronóstico: se toma la fecha de finalización y se normaliza al primer día del mes
    forecast_start = fecha_fin.replace(day=1) + pd.DateOffset(months=1)
    forecast_index = pd.date_range(start=forecast_start, periods=n_pronosticos, freq='MS')

    # Obtener pronóstico de 12 meses
    forecast_obj = resultado.get_forecast(steps=n_pronosticos)
    prediccion_log = forecast_obj.predicted_mean
    prediccion = np.expm1(prediccion_log)  # Invertir la transformación log1p

    # Crear DataFrame con la predicción para este ingrediente
    df_pred = pd.DataFrame({
         "Ingrediente": ing,
         "Fecha": forecast_index,
         "Consumo Predicho": prediccion.values
    })
    predicciones_list.append(df_pred)
    
    print(f"\n--- Pronóstico para {ing} ---")
    print(df_pred)

# Importar para graficar
import matplotlib.pyplot as plt
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf

for ing, vector in vectores_consumo.items():
    # Crear índice de fechas para la serie histórica
    index_historico = pd.date_range(start=fecha_inicio, periods=len(vector), freq='MS')
    serie = pd.Series(data=vector, index=index_historico)
    
    print(f"\n--- Serie histórica para {ing} ---")
    print(serie.head())
    print("...")
    print(serie.tail())
    ultimo_original = serie.index[-1]
    
    # Transformación logarítmica para estabilizar la varianza
    serie_log = np.log1p(serie)
    
    # Ajustar modelo SARIMAX sobre la serie
    try:
        modelo = SARIMAX(
            serie_log,
            order=(1, 1, 1),
            seasonal_order=(1, 1, 1, 12),
            enforce_stationarity=False,
            enforce_invertibility=False
        )
        resultado = modelo.fit(disp=False)
        print(f"Modelo ajustado correctamente para {ing}.")
    except Exception as e:
        print(f"Error al ajustar el modelo para {ing}: {e}")
        continue
    
    # --- Fase de Prueba: Graficar Diagnósticos ---
    
    # 1. Graficar la serie histórica y la predicción (ajustada) para visualizar el comportamiento
    forecast_start = fecha_fin.replace(day=1) + pd.DateOffset(months=1)
    forecast_index = pd.date_range(start=forecast_start, periods=n_pronosticos, freq='MS')
    forecast_obj = resultado.get_forecast(steps=n_pronosticos)
    prediccion_log = forecast_obj.predicted_mean
    prediccion = np.expm1(prediccion_log)  # Invertir log1p
    
    plt.figure(figsize=(10, 4))
    plt.plot(serie.index, serie, label='Histórico')
    plt.plot(forecast_index, prediccion, label='Predicción', marker='o')
    plt.title(f"Serie histórica y predicción para {ing}")
    plt.xlabel("Fecha")
    plt.ylabel("Consumo")
    plt.legend()
    plt.tight_layout()
    plt.show()
    
    # 2. Extraer y graficar los residuos del modelo
    residuos = resultado.resid
    plt.figure(figsize=(12, 4))
    plt.plot(residuos, marker='o')
    plt.title(f"Residuos del modelo SARIMAX para {ing}")
    plt.xlabel("Fecha")
    plt.ylabel("Residuos")
    plt.tight_layout()
    plt.show()
    
    # 3. Graficar el ACF y PACF de los residuos
    fig, ax = plt.subplots(1, 2, figsize=(12, 4))
    plot_acf(residuos, ax=ax[0], title=f'ACF de residuos - {ing}')
    plot_pacf(residuos, ax=ax[1], title=f'PACF de residuos - {ing}', method='ywm')
    plt.tight_layout()
    plt.show()
    
    # Continuamos con la generación del DataFrame de predicción...
    df_pred = pd.DataFrame({
         "Ingrediente": ing,
         "Fecha": forecast_index,
         "Consumo Predicho": prediccion.values
    })
    predicciones_list.append(df_pred)
    
    print(f"\n--- Pronóstico para {ing} ---")
    print(df_pred)

# Concatenar todas las predicciones en un solo DataFrame
df_predicciones = pd.concat(predicciones_list, ignore_index=True)
df_predicciones["Año"] = df_predicciones["Fecha"].dt.year
df_predicciones["Mes"] = df_predicciones["Fecha"].dt.month

# Crear la tabla pivot: filas = (Ingrediente, Año) y columnas = Mes
pivot_pred = df_predicciones.pivot(index=["Ingrediente", "Año"], columns="Mes", values="Consumo Predicho")
pivot_pred = pivot_pred.reindex(columns=range(1, 13), fill_value=0)
pivot_pred = pivot_pred.astype(float)
pivot_pred.rename(columns=meses_abreviados, inplace=True)

# Redondear los valores a 2 decimales
pivot_pred = pivot_pred.round(2)

# Exportar la tabla de predicciones a Excel con autoajuste de columnas
ruta_pred = "prediccion_simple.xlsx"
with pd.ExcelWriter(ruta_pred, engine="openpyxl") as writer:
    pivot_pred.to_excel(writer, sheet_name="Predicciones", index=True)
    worksheet = writer.sheets["Predicciones"]
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[col_letter].width = adjusted_width
print(f"Archivo de predicciones generado: {ruta_pred}")

