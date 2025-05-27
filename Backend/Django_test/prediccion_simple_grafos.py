import os
import django
import pandas as pd
from django.db.models import Sum, F, Min, Max
import warnings
from itertools import product  
from openpyxl.utils import get_column_letter
import numpy as np
import matplotlib.pyplot as plt
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tools.sm_exceptions import ConvergenceWarning
from sklearn.metrics import mean_absolute_error, mean_squared_error

# Filtramos warnings generales y de convergencia de statsmodels
warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", category=ConvergenceWarning)

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "hotel_management.settings")
django.setup()

from gestion.models import OrdenElementos, CocktailIngredientes, Ingredientes

def generar_consumos_historicos_excel():
    # Obtener fecha de inicio (primer registro) y de finalización (último registro)
    fecha_inicio = OrdenElementos.objects.aggregate(primera_fecha=Min("fechaorden"))["primera_fecha"]
    fecha_fin = OrdenElementos.objects.aggregate(ultima_fecha=Max("fechaorden"))["ultima_fecha"]

    if not fecha_inicio or not fecha_fin:
        print("No hay registros en la base de datos para calcular consumos.")
        return

    print(f"Fecha de inicio detectada automáticamente: {fecha_inicio}")
    print(f"Fecha de finalización detectada automáticamente: {fecha_fin}")

    consumo_total = {}

    # Cálculo de consumo para registros directos (no cocktail)
    datos_directos = (
        OrdenElementos.objects
        .filter(escocktail=False, fechaorden__range=[fecha_inicio, fecha_fin])
        .annotate(
            año=F('fechaorden__year'),
            mes=F('fechaorden__month'),
            ingrediente_nombre=F('id_ingredientes__nombre')
        )
        .values('ingrediente_nombre', 'año', 'mes')
        .annotate(total_litros=Sum(F('cantidad') * F('id_ingredientes__litrosporunidad')))
    )

    for dato in datos_directos:
        key = (dato['ingrediente_nombre'], dato['año'], dato['mes'])
        consumo_total[key] = consumo_total.get(key, 0) + dato['total_litros']

    # Cálculo de consumo para registros de cocktails
    datos_cocktails = OrdenElementos.objects.filter(
        escocktail=True,
        fechaorden__range=[fecha_inicio, fecha_fin]
    )
    for cocktail in datos_cocktails:
        receta = CocktailIngredientes.objects.filter(
            id_cocktail=cocktail.id_cocktail,
            activo=True
        )
        for ingrediente in receta:
            key = (ingrediente.id_ingredientes.nombre, cocktail.fechaorden.year, cocktail.fechaorden.month)
            cantidad_usada = cocktail.cantidad * ingrediente.cantidad
            consumo_total[key] = consumo_total.get(key, 0) + cantidad_usada

    # Obtener todos los ingredientes existentes en la base de datos
    ingredientes = Ingredientes.objects.values_list("nombre", flat=True).distinct()
    fecha_fin_datetime = pd.to_datetime(fecha_fin)
    años = range(fecha_inicio.year, fecha_fin_datetime.year + 1)
    meses = range(1, 13)

    combinaciones = []
    for año in años:
        if año == fecha_fin_datetime.year:
            meses_hasta_fin = range(1, fecha_fin_datetime.month + 1)
            combinaciones.extend(product(ingredientes, [año], meses_hasta_fin))
        else:
            combinaciones.extend(product(ingredientes, [año], meses))

    datos_consumo = [
        {
            "Ingrediente": ingrediente,
            "Año": año,
            "Mes": mes,
            "Consumo Total": consumo_total.get((ingrediente, año, mes), 0)
        }
        for ingrediente, año, mes in combinaciones
    ]

    df = pd.DataFrame(datos_consumo)
    df.sort_values(by=['Ingrediente', 'Año', 'Mes'], inplace=True)

    # Crear la tabla pivot: filas = (Ingrediente, Año) y columnas = Mes (1 a 12)
    pivot_table = df.pivot(index=["Ingrediente", "Año"], columns="Mes", values="Consumo Total")
    pivot_table = pivot_table.reindex(columns=range(1, 13), fill_value=0)
    pivot_table = pivot_table.astype(float)

    # Renombrar columnas a abreviaturas de 3 letras
    meses_abreviados = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
        5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
    }
    pivot_table.rename(columns=meses_abreviados, inplace=True)

    # --- Generación de vectores de consumos históricos ---
    periodos = []
    current_year = fecha_inicio.year
    current_month = fecha_inicio.month
    while current_year < fecha_fin.year or (current_year == fecha_fin.year and current_month <= fecha_fin.month):
        periodos.append((current_year, current_month))
        if current_month == 12:
            current_year += 1
            current_month = 1
        else:
            current_month += 1

    vectores_consumo = {}
    for ing in ingredientes:
        vector_consumo = [float(consumo_total.get((ing, año, mes), 0)) for (año, mes) in periodos]
        vectores_consumo[ing] = vector_consumo
        print(f"{ing}: {vector_consumo}")    

    # Exportar la tabla pivot a Excel con autoajuste de columnas
    ruta_archivo = "consumo_historico.xlsx"
    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        pivot_table.to_excel(writer, sheet_name="Consumos", index=True)
        worksheet = writer.sheets["Consumos"]
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    print(f"Archivo generado: {ruta_archivo}")
    return fecha_inicio, fecha_fin, vectores_consumo

# Ejecutar la función histórica y obtener vectores
fecha_inicio, fecha_fin, vectores_consumo = generar_consumos_historicos_excel()

# ========================================================================================
# Fase 2: Modelado, selección de parámetros (usando AIC y BIC) y diagnóstico gráfico
n_pronosticos = 12
predicciones_list = []
meses_abreviados = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
}

for ing, vector in vectores_consumo.items():
    # Crear índice de fechas para la serie histórica
    index_historico = pd.date_range(start=fecha_inicio, periods=len(vector), freq='MS')
    serie = pd.Series(data=vector, index=index_historico)
    
    print(f"\n--- Serie histórica para {ing} ---")
    print(serie.head())
    print("...")
    print(serie.tail())
    
    # Transformación logarítmica para estabilizar la varianza
    serie_log = np.log1p(serie)
    
    # ---------------------------
    # Fase de comparación de modelos usando AIC y BIC (Grid Search)
    # ---------------------------
    best_aic = np.inf
    best_bic = np.inf
    best_order = None
    best_seasonal_order = None
    best_model = None

    # Grid search: probar combinaciones de parámetros (p, d, q) y estacionales (P, D, Q, s=12)
    for p in range(0, 3):
        for d in range(0, 2):
            for q in range(0, 3):
                for P in range(0, 2):
                    for D in range(0, 2):
                        for Q in range(0, 2):
                            try:
                                mod = SARIMAX(serie_log, 
                                              order=(p, d, q), 
                                              seasonal_order=(P, D, Q, 12),
                                              enforce_stationarity=False, 
                                              enforce_invertibility=False)
                                res = mod.fit(disp=False, maxiter=300, method='lbfgs')
                                if (res.aic < best_aic) and (res.bic < best_bic):
                                    best_aic = res.aic
                                    best_bic = res.bic
                                    best_order = (p, d, q)
                                    best_seasonal_order = (P, D, Q, 12)
                                    best_model = res
                            except Exception:
                                continue

    print(f"Mejor modelo para {ing}: order={best_order}, seasonal_order={best_seasonal_order}, AIC={best_aic:.2f}, BIC={best_bic:.2f}")
    resultado = best_model

    # ---------------------------
    # Fase de validación cruzada (walk-forward forecasting) para evaluar generalización
    # ---------------------------
    # Usaremos el 80% de los datos para entrenamiento y el 20% restante para prueba.
    split_ratio = 0.8
    n_total = len(serie_log)
    n_train = int(np.floor(n_total * split_ratio))
    train_series = serie_log.iloc[:n_train]
    test_series = serie_log.iloc[n_train:]
    walk_preds = []
    history = list(train_series)
    for t in range(len(test_series)):
        try:
            model_cv = SARIMAX(history, order=best_order, seasonal_order=best_seasonal_order,
                                enforce_stationarity=False, enforce_invertibility=False)
            model_cv_fit = model_cv.fit(disp=False, maxiter=300, method='lbfgs')
        except Exception:
            break
        forecast_log_cv = model_cv_fit.get_forecast(steps=1).predicted_mean
        forecast_cv = np.expm1(forecast_log_cv)[0]
        walk_preds.append(forecast_cv)
        history.append(test_series.iloc[t])
    # Convertir los valores reales a dominio original (invirtiendo log1p)
    actuals_cv = np.expm1(test_series.values)
    mae_cv = mean_absolute_error(actuals_cv, walk_preds)
    rmse_cv = np.sqrt(mean_squared_error(actuals_cv, walk_preds))
    mape_cv = np.mean(np.abs((actuals_cv - walk_preds) / actuals_cv)) * 100
    print(f"Validación cruzada para {ing}: MAE={mae_cv:.2f}, RMSE={rmse_cv:.2f}, MAPE={mape_cv:.2f}%")
    
    # Gráfico de validación cruzada: comparando valores reales vs walk-forward forecast
    plt.figure(figsize=(8,4))
    plt.plot(test_series.index, actuals_cv, marker='o', label="Valores reales")
    plt.plot(test_series.index, walk_preds, marker='o', linestyle='--', color='red', label="Walk-forward Forecast")
    plt.title(f"Validación cruzada - {ing}")
    plt.xlabel("Fecha")
    plt.ylabel("Consumo")
    plt.legend()
    plt.tight_layout()
    plt.show()

    # ---------------------------
    # Fase de diagnóstico gráfico: pronóstico futuro y análisis de residuos
    # ---------------------------
    forecast_start = fecha_fin.replace(day=1) + pd.DateOffset(months=1)
    forecast_index = pd.date_range(start=forecast_start, periods=n_pronosticos, freq='MS')

    forecast_obj = resultado.get_forecast(steps=n_pronosticos)
    prediccion_log = forecast_obj.predicted_mean
    prediccion = np.expm1(prediccion_log)

    # Diagnóstico gráfico de residuos
    residuals = resultado.resid
    max_lags = min(20, int(np.floor(len(residuals) * 0.5)) - 1)
    if max_lags < 1:
        max_lags = 1

    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    axes[0, 0].plot(serie, marker='o', linestyle='-', label="Serie histórica")
    axes[0, 0].set_title(f"Serie histórica - {ing}")
    axes[0, 0].legend()
    axes[0, 1].plot(serie, marker='o', linestyle='-', label="Serie histórica")
    axes[0, 1].plot(forecast_index, prediccion, marker='o', linestyle='--', color='red', label="Pronóstico")
    axes[0, 1].set_title(f"Pronóstico - {ing}")
    axes[0, 1].legend()
    plot_acf(residuals, ax=axes[1, 0], lags=max_lags)
    axes[1, 0].set_title(f"ACF de residuos - {ing}")
    plot_pacf(residuals, ax=axes[1, 1], lags=max_lags)
    axes[1, 1].set_title(f"PACF de residuos - {ing}")
    plt.tight_layout()
    plt.show()

    # Crear DataFrame con la predicción para este ingrediente
    df_pred = pd.DataFrame({
         "Ingrediente": ing,
         "Fecha": forecast_index,
         "Consumo Predicho": prediccion.values
    })
    predicciones_list.append(df_pred)
    
    print(f"\n--- Pronóstico para {ing} ---")
    print(df_pred)

# Concatenar todas las predicciones en un solo DataFrame y generar tabla pivot
df_predicciones = pd.concat(predicciones_list, ignore_index=True)
df_predicciones["Año"] = df_predicciones["Fecha"].dt.year
df_predicciones["Mes"] = df_predicciones["Fecha"].dt.month
pivot_pred = df_predicciones.pivot(index=["Ingrediente", "Año"], columns="Mes", values="Consumo Predicho")
pivot_pred = pivot_pred.reindex(columns=range(1, 13), fill_value=0).astype(float)
pivot_pred.rename(columns=meses_abreviados, inplace=True)
pivot_pred = pivot_pred.round(2)

ruta_pred = "prediccion_simple.xlsx"
with pd.ExcelWriter(ruta_pred, engine="openpyxl") as writer:
    pivot_pred.to_excel(writer, sheet_name="Predicciones", index=True)
    worksheet = writer.sheets["Predicciones"]
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[col_letter].width = max_length + 2
print(f"Archivo de predicciones generado: {ruta_pred}")
