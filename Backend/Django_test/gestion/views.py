from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import JsonResponse, FileResponse
from django.views import View
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.exceptions import ValidationError
from django.db.models import Count, Case, When, CharField, Sum, F, Min, Exists, OuterRef, Subquery
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils.timezone import now
from django.db import transaction, IntegrityError
import pandas as pd
import numpy as np
import logging

from decimal import Decimal
from itertools import product
from statsmodels.tsa.statespace.sarimax import SARIMAX
import datetime
import tempfile
from gestion.models import Ingredientes, OrdenElementos, CocktailIngredientes
from .models import (Ingredientes, Cocktail, AsignacionesHuespedes,
                    Ordenes, OrdenElementos, Empleados, Bares,
                    CocktailIngredientes, Huesped, Habitaciones,
                    AsignacionesHuespedes, MovimientosStock, RegistroOrdenes,
                    RegistroMovimientos, RegistroEstancias)
from .serializers import (CrearOrdenSerializer, CrearOrdenElementoSerializer,
                        BaresSerializer,IngredientesSerializer,
                        CocktailSerializer , LoginSerializer,
                        AsignacionesHuespedesSerializer, OrdenesSerializer,
                        OrdenElementosSerializer, CocktailIngredientesSerializer,
                        CrearCocktailIngredientesSerializer, UpdateCocktailIngredienteSerializer,
                        CrearIngredienteSerializer, CrearCocktailSerializer,
                        HabitacionDisponibleSerializer, HuespedSerializer,
                        AsignacionesHuespedesDetailSerializer, CrearAsignacionSerializer,
                        HabitacionesSerializer, EmpleadoSerializer,
                        MovimientoStockSerializer, CrearOrdenConElementosSerializer,
                        RegistroOrdenesSerializer, RegistroMovimientosSerializer,
                        RegistroEstanciasSerializer)


class AsignacionesActivasView(APIView):
    def get(self, request):
        try:
            asignaciones = AsignacionesHuespedes.objects.filter(enhotel=True)
            serializer = AsignacionesHuespedesSerializer(asignaciones, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class OrdenElementosView(APIView):
    def get(self, request, id_orden):
        try:
            elementos = OrdenElementos.objects.filter(id_orden=id_orden).select_related('id_cocktail', 'id_ingredientes')
            serializer = OrdenElementosSerializer(elementos, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except OrdenElementos.DoesNotExist:
            return Response({"error": "No se encontraron elementos para esta orden."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": f"Error interno: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class BaresListView(APIView):
    def get(self, request):
        try:
            bares = Bares.objects.all()
            serializer = BaresSerializer(bares, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class LoginView(APIView):
    def post(self, request, format=None):
        password = request.data.get("password")
        bar_id = request.data.get("bar")

        if not password or not bar_id:
            return Response({"error": "Empleado ID y Bar son requeridos"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            empleado = Empleados.objects.get(password=password)

            if not empleado.activo:
                return Response(
                    {"error": "El empleado está inactivo y no puede iniciar sesión."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            serializer = LoginSerializer(empleado)
            return Response({"empleado": serializer.data}, status=status.HTTP_200_OK)
        
        except Empleados.DoesNotExist:
            return Response(
                {"error": "Empleado no encontrado con las credenciales proporcionadas."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Error interno del servidor: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AsignacionesPorHabitacionView(APIView):
    def get(self, request, habitacion):
        try:
            asignaciones = AsignacionesHuespedes.objects.filter(
                numerohabitacion__numerohabitacion=habitacion,
                enhotel=True
            ).select_related('id_huesped', 'numerohabitacion')

            if asignaciones.exists():
                serializer = AsignacionesHuespedesSerializer(asignaciones, many=True)
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(
                    {"error": f"No hay asignaciones activas para la habitación {habitacion}."},
                    status=status.HTTP_404_NOT_FOUND,
                )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class IngredientesListView(APIView):
    def get(self, request):
        try:
            ingredientes = Ingredientes.objects.all()
            serializer = IngredientesSerializer(ingredientes, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CocktailListView(APIView):
    def get(self, request):
        try:
            cocktails = Cocktail.objects.all()
            serializer = CocktailSerializer(cocktails, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateCocktailView(APIView):
    def patch(self, request, id_cocktail):
        try:
            cocktail = Cocktail.objects.get(id_cocktail=id_cocktail)
            serializer = CocktailSerializer(cocktail, data=request.data, partial=True)

            if serializer.is_valid():
                serializer.save()
                return Response({
                    "mensaje": "Cóctel actualizado correctamente.",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )
        except IntegrityError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST )
        except Cocktail.DoesNotExist:
            return Response({
                "error": f"No se encontró un cóctel con ID {id_cocktail}"
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "error": f"Error al actualizar el cóctel: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CrearOrdenView(APIView):
    def post(self, request):
        serializer = CrearOrdenSerializer(data=request.data)
        if serializer.is_valid():
            try:
                orden = serializer.save()
                return Response(
                    {
                        "mensaje": "Orden creada exitosamente.",
                        "id_orden": orden.id_orden,
                    },
                    status=status.HTTP_201_CREATED,
                )
            except IntegrityError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response(
                    {"error": f"Error al crear la orden: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CrearOrdenElementoView(APIView):
    def post(self, request):
        try:
            serializer = CrearOrdenElementoSerializer(data=request.data)
            if serializer.is_valid():
                try:
                    id_orden = serializer.validated_data["id_orden"]
                    cantidad = serializer.validated_data["cantidad"]
                    id_cocktail = serializer.validated_data.get("id_cocktail")
                    id_ingredientes = serializer.validated_data.get("id_ingredientes")

                    if not id_cocktail and not id_ingredientes:
                        return Response(
                            {"error": "Debe proporcionar id_cocktail o id_ingredientes."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    escocktail = bool(id_cocktail)
                    elemento = OrdenElementos.objects.create(
                        id_orden_id=id_orden,
                        id_cocktail_id=id_cocktail if escocktail else None,
                        id_ingredientes_id=id_ingredientes if not escocktail else None,
                        cantidad=cantidad,
                        escocktail=escocktail,
                        preciototal=0,
                    )
                    return Response(
                        {
                            "mensaje": "Elemento de la orden creado exitosamente.",
                            "id_elemento": elemento.id_elemento,
                        },
                        status=status.HTTP_201_CREATED,
                    )
                except ValidationError as ve:
                    return Response({"error": ve.messages}, status=status.HTTP_400_BAD_REQUEST)
                except IntegrityError as e:
                    return Response({"error": f"Error de integridad: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
                except Exception as e:
                    return Response({"error": f"Error interno: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Error inesperado: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


logger = logging.getLogger('movimientosstock')

class CrearOrdenConElementosView(APIView):
    def post(self, request):
        serializer = CrearOrdenConElementosSerializer(data=request.data)
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    orden = serializer.save()
                    self.actualizar_stock(orden)
                logger.info(
                    f"Orden creada exitosamente: ID {orden.id_orden} "
                    f"para Asignación {orden.id_asignacion.id_asignacion} "
                    f"en Bar {orden.id_bar.nombre} por Empleado {orden.id_empleado.nombre}. "
                    f"Precio Total: {orden.preciototal}€"
                )
                return Response(
                    {
                        "mensaje": "Orden y elementos creados exitosamente.",
                        "id_orden": orden.id_orden,
                    },
                    status=status.HTTP_201_CREATED,
                )
            except ValidationError as ve:
                logger.error(f"ValidationError al crear orden: {ve.messages}")
                return Response({"error": ve.messages}, status=status.HTTP_400_BAD_REQUEST)
            except IntegrityError as e:
                logger.error(f"IntegrityError al crear orden: {str(e)}")
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                logger.exception("Error inesperado al crear orden.")
                return Response(
                    {"error": f"Error al crear la orden: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        else:
            logger.warning(f"Datos inválidos para crear orden: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def actualizar_stock(self, orden):
        for elemento in orden.ordenelementos_set.all():
            if elemento.escocktail:
                receta = CocktailIngredientes.objects.filter(
                    id_cocktail=elemento.id_cocktail,
                    activo=True
                )
                if not receta.exists():
                    logger.error(
                        f"El cóctel '{elemento.id_cocktail.nombre}' no tiene ingredientes registrados."
                    )
                    raise ValidationError(
                        f"El cóctel '{elemento.id_cocktail.nombre}' no tiene ingredientes registrados."
                    )
                for ingrediente in receta:
                    try:
                        stock_ingrediente = Ingredientes.objects.select_for_update().get(
                            id_ingredientes=ingrediente.id_ingredientes.id_ingredientes
                        )
                    except Ingredientes.DoesNotExist:
                        logger.error(
                            f"Ingrediente con ID {ingrediente.id_ingredientes.id_ingredientes} no encontrado."
                        )
                        raise ValidationError(
                            f"Ingrediente con ID {ingrediente.id_ingredientes.id_ingredientes} no encontrado."
                        )
                    cantidad_total = (
                        Decimal(ingrediente.cantidad) * Decimal(elemento.cantidad)
                    ).quantize(Decimal("0.001"))
                    if stock_ingrediente.cantidadactual < cantidad_total:
                        logger.error(
                            f"Stock insuficiente para '{stock_ingrediente.nombre}'. "
                            f"Requerido: {cantidad_total} L, Disponible: {stock_ingrediente.cantidadactual} L."
                        )
                        raise ValidationError(
                            f"Stock insuficiente para '{stock_ingrediente.nombre}'. "
                            f"Requerido: {cantidad_total} L, Disponible: {stock_ingrediente.cantidadactual} L."
                        )
                    stock_ingrediente.cantidadactual -= cantidad_total
                    stock_ingrediente.save()
                    MovimientosStock.objects.create(
                        id_ingredientes=stock_ingrediente,
                        tipomovimiento="Consumo (orden)",
                        cantidad=cantidad_total
                    )
                    logger.info(
                        f"Descontado {cantidad_total} L de '{stock_ingrediente.nombre}'. "
                        f"Stock restante: {stock_ingrediente.cantidadactual} L."
                    )
            else:
                try:
                    ingrediente = Ingredientes.objects.select_for_update().get(
                        id_ingredientes=elemento.id_ingredientes.id_ingredientes
                    )
                except Ingredientes.DoesNotExist:
                    logger.error(
                        f"Ingrediente con ID {elemento.id_ingredientes.id_ingredientes} no encontrado."
                    )
                    raise ValidationError(
                        f"Ingrediente con ID {elemento.id_ingredientes.id_ingredientes} no encontrado."
                    )
                cantidad_a_descontar = (
                    Decimal(ingrediente.litrosporunidad) * Decimal(elemento.cantidad)
                ).quantize(Decimal("0.001"))
                if ingrediente.cantidadactual < cantidad_a_descontar:
                    logger.error(
                        f"Stock insuficiente para '{ingrediente.nombre}'. "
                        f"Requerido: {cantidad_a_descontar} L, Disponible: {ingrediente.cantidadactual} L."
                    )
                    raise ValidationError(
                        f"Stock insuficiente para '{ingrediente.nombre}'. "
                        f"Requerido: {cantidad_a_descontar} L, Disponible: {ingrediente.cantidadactual} L."
                    )
                ingrediente.cantidadactual -= cantidad_a_descontar
                ingrediente.save()
                MovimientosStock.objects.create(
                    id_ingredientes=ingrediente,
                    tipomovimiento="Consumo (orden)",
                    cantidad=cantidad_a_descontar
                )
                logger.info(
                    f"Descontado {cantidad_a_descontar} L de '{ingrediente.nombre}'. "
                    f"Stock restante: {ingrediente.cantidadactual} L."
                )


class AgregarElementoOrdenView(APIView):
    def post(self, request, id_orden):
        try:
            with transaction.atomic():
                orden = Ordenes.objects.select_for_update().get(id_orden=id_orden, actividad=True)
                
                elementos_data = request.data 
                if not isinstance(elementos_data, list):
                    return Response({"error": "Se espera una lista de elementos."}, status=status.HTTP_400_BAD_REQUEST)
                
                elementos_creados = []
                
                for elemento_data in elementos_data:
                    serializer = CrearOrdenElementoSerializer(data=elemento_data)
                    if serializer.is_valid():
                        id_cocktail = serializer.validated_data.get("id_cocktail")
                        id_ingredientes = serializer.validated_data.get("id_ingredientes")
                        cantidad = serializer.validated_data["cantidad"]
                        
                        escocktail = bool(id_cocktail)
                        
                        elemento = OrdenElementos.objects.create(
                            id_orden=orden,
                            id_cocktail_id=id_cocktail if escocktail else None,
                            id_ingredientes_id=id_ingredientes if not escocktail else None,
                            cantidad=cantidad,
                            escocktail=escocktail,
                        )
                        
                        self.actualizar_stock(elemento)

                        if elemento.escocktail and elemento.id_cocktail:
                            nombre = elemento.id_cocktail.nombre
                        elif not elemento.escocktail and elemento.id_ingredientes:
                            nombre = elemento.id_ingredientes.nombre
                        else:
                            nombre = "N/A"
                        
                        elementos_creados.append({
                            "id_elemento": elemento.id_elemento,
                            "nombre": nombre,
                            "cantidad": elemento.cantidad,
                            "preciototal": elemento.preciototal,
                        })
                    else:
                        raise ValidationError(serializer.errors)
                
                return Response(
                    {
                        "mensaje": "Elementos agregados a la orden exitosamente.",
                        "elementos_agregados": elementos_creados,
                    },
                    status=status.HTTP_201_CREATED,
                )
        except Ordenes.DoesNotExist:
            logger.error(f"Orden con ID {id_orden} no encontrada o inactiva.")
            return Response(
                {"error": f"Orden con ID {id_orden} no encontrada o inactiva."},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValidationError as ve:
            logger.error(f"ValidationError al agregar elementos a la orden: {ve.messages}")
            return Response({"error": ve.messages}, status=status.HTTP_400_BAD_REQUEST)
        except IntegrityError as e:
            logger.error(f"IntegrityError al agregar elementos a la orden: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Ingredientes.DoesNotExist as e:
            logger.error(f"Ingrediente relacionado no encontrado: {str(e)}")
            return Response(
                {"error": "Ingrediente relacionado no encontrado."},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception("Error inesperado al agregar elementos a la orden.")
            return Response(
                {"error": f"Error al agregar los elementos: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def actualizar_stock(self, elemento):
        if elemento.escocktail:
            receta = CocktailIngredientes.objects.filter(
                id_cocktail=elemento.id_cocktail,
                activo=True
            )
            if not receta.exists():
                logger.error(f"El cóctel '{elemento.id_cocktail.nombre}' no tiene ingredientes registrados.")
                raise ValidationError(f"El cóctel '{elemento.id_cocktail.nombre}' no tiene ingredientes registrados.")
            
            for ingrediente_receta in receta:
                try:
                    stock_ingrediente = Ingredientes.objects.select_for_update().get(
                        id_ingredientes=ingrediente_receta.id_ingredientes.id_ingredientes
                    )
                except Ingredientes.DoesNotExist:
                    logger.error(f"Ingrediente con ID {ingrediente_receta.id_ingredientes.id_ingredientes} no encontrado.")
                    raise ValidationError(f"Ingrediente con ID {ingrediente_receta.id_ingredientes.id_ingredientes} no encontrado.")
                
                cantidad_total = (
                    Decimal(ingrediente_receta.cantidad) * Decimal(elemento.cantidad)
                ).quantize(Decimal("0.001"))
                
                if stock_ingrediente.cantidadactual < cantidad_total:
                    logger.error(
                        f"Stock insuficiente para '{stock_ingrediente.nombre}'. "
                        f"Requerido: {cantidad_total} L, Disponible: {stock_ingrediente.cantidadactual} L."
                    )
                    raise ValidationError(
                        f"Stock insuficiente para '{stock_ingrediente.nombre}'. "
                        f"Requerido: {cantidad_total} L, Disponible: {stock_ingrediente.cantidadactual} L."
                    )
                
                stock_ingrediente.cantidadactual -= cantidad_total
                stock_ingrediente.save()
                
                MovimientosStock.objects.create(
                    id_ingredientes=stock_ingrediente,
                    tipomovimiento="Consumo (orden)",
                    cantidad=cantidad_total
                )
                
                logger.info(
                    f"Descontado {cantidad_total} L de '{stock_ingrediente.nombre}'. "
                    f"Stock restante: {stock_ingrediente.cantidadactual} L."
                )
        else:
            try:
                ingrediente = Ingredientes.objects.select_for_update().get(
                    id_ingredientes=elemento.id_ingredientes.id_ingredientes
                )
            except Ingredientes.DoesNotExist:
                logger.error(f"Ingrediente con ID {elemento.id_ingredientes.id_ingredientes} no encontrado.")
                raise ValidationError(f"Ingrediente con ID {elemento.id_ingredientes.id_ingredientes} no encontrado.")
            
            cantidad_a_descontar = (
                Decimal(ingrediente.litrosporunidad) * Decimal(elemento.cantidad)
            ).quantize(Decimal("0.001"))
            
            if ingrediente.cantidadactual < cantidad_a_descontar:
                logger.error(
                    f"Stock insuficiente para '{ingrediente.nombre}'. "
                    f"Requerido: {cantidad_a_descontar} L, Disponible: {ingrediente.cantidadactual} L."
                )
                raise ValidationError(
                    f"Stock insuficiente para '{ingrediente.nombre}'. "
                    f"Requerido: {cantidad_a_descontar} L, Disponible: {ingrediente.cantidadactual} L."
                )
            
            ingrediente.cantidadactual -= cantidad_a_descontar
            ingrediente.save()
            
            MovimientosStock.objects.create(
                id_ingredientes=ingrediente,
                tipomovimiento="Consumo (orden)",
                cantidad=cantidad_a_descontar
            )
            
            logger.info(
                f"Descontado {cantidad_a_descontar} L de '{ingrediente.nombre}'. "
                f"Stock restante: {ingrediente.cantidadactual} L."
            )


class EliminarOrdenElementoView(APIView):
    def delete(self, request, id_orden, id_elemento):
        try:
            with transaction.atomic():
                elemento = OrdenElementos.objects.select_for_update().get(
                    id_orden=id_orden,
                    id_elemento=id_elemento
                )

                if elemento.escocktail:
                    receta = CocktailIngredientes.objects.filter(
                        id_cocktail=elemento.id_cocktail,
                        activo=True
                    )
                    if not receta.exists():
                        logger.error(
                            f"El cóctel '{elemento.id_cocktail.nombre}' no tiene ingredientes registrados."
                        )
                        raise ValidationError(
                            f"El cóctel '{elemento.id_cocktail.nombre}' no tiene ingredientes registrados."
                        )
                    
                    for ingrediente_receta in receta:
                        ingrediente = Ingredientes.objects.select_for_update().get(
                            id_ingredientes=ingrediente_receta.id_ingredientes.id_ingredientes
                        )
                        cantidad_total = (
                            Decimal(ingrediente_receta.cantidad) * Decimal(elemento.cantidad)
                        ).quantize(Decimal("0.001"))
                        
                        ingrediente.cantidadactual += cantidad_total
                        ingrediente.save()
                        
                        MovimientosStock.objects.create(
                            id_ingredientes=ingrediente,
                            tipomovimiento="Cancelacion (orden)",
                            cantidad=cantidad_total
                        )
                        
                        logger.info(
                            f"Reabastecido {cantidad_total} L de '{ingrediente.nombre}'. "
                            f"Stock total: {ingrediente.cantidadactual} L."
                        )
                else:
                    ingrediente = Ingredientes.objects.select_for_update().get(
                        id_ingredientes=elemento.id_ingredientes.id_ingredientes
                    )
                    cantidad_a_reabastecer = (
                        Decimal(ingrediente.litrosporunidad) * Decimal(elemento.cantidad)
                    ).quantize(Decimal("0.001"))
                    
                    ingrediente.cantidadactual += cantidad_a_reabastecer
                    ingrediente.save()
                    
                    MovimientosStock.objects.create(
                        id_ingredientes=ingrediente,
                        tipomovimiento="Cancelacion (orden)",
                        cantidad=cantidad_a_reabastecer
                    )
                    
                    logger.info(
                        f"Reabastecido {cantidad_a_reabastecer} L de '{ingrediente.nombre}'. "
                        f"Stock total: {ingrediente.cantidadactual} L."
                    )
                
                elemento.delete()
                logger.info(
                    f"Elemento eliminado: ID {id_elemento} de Orden {id_orden}."
                )
                
            return Response(
                {"mensaje": "Elemento eliminado y stock reabastecido correctamente."},
                status=status.HTTP_200_OK
            )
        except OrdenElementos.DoesNotExist:
            logger.error(f"Elemento con ID {id_elemento} en Orden {id_orden} no encontrado.")
            return Response(
                {"error": f"Elemento con ID {id_elemento} en Orden {id_orden} no encontrado."},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValidationError as ve:
            logger.error(f"ValidationError al eliminar elemento: {ve.messages}")
            return Response({"error": ve.messages}, status=status.HTTP_400_BAD_REQUEST)
        except Ingredientes.DoesNotExist as e:
            logger.error(f"Ingrediente relacionado no encontrado: {str(e)}")
            return Response(
                {"error": "Ingrediente relacionado no encontrado."},
                status=status.HTTP_400_BAD_REQUEST
            )
        except IntegrityError as e:
            logger.error(f"IntegrityError al eliminar elemento: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception("Error inesperado al eliminar elemento de la orden.")
            return Response(
                {"error": f"Error al eliminar el elemento: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class OrdenesActivasView(APIView):
    def get(self, request):
        bar_id = request.query_params.get("bar_id")
        if not bar_id:
            return Response({"error": "El ID del bar es requerido"}, status=400)
        try:
            ordenes = Ordenes.objects.filter(id_bar=bar_id, actividad=True).select_related("id_asignacion", "id_empleado", "id_bar")
            serializer = OrdenesSerializer(ordenes, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"Error al obtener órdenes activas: {str(e)}"}, status=500)


class CocktailIngredientesList(APIView):
    def get(self, request, cocktail_id, format=None):
        try:
            cocktail = Cocktail.objects.get(id_cocktail=cocktail_id)
            ingredientes_qs = CocktailIngredientes.objects.filter(id_cocktail=cocktail_id, activo=True)
            ing_serializer = CocktailIngredientesSerializer(ingredientes_qs, many=True)

            data = {
                "id_cocktail": cocktail.id_cocktail,
                "nombre": cocktail.nombre,
                "version": cocktail.version,
                "precioporunidad": float(cocktail.precioporunidad),
                "tienereceta": cocktail.tienereceta,
                "activo": cocktail.activo, 
                "ingredientes": ing_serializer.data
            }
            return Response(data, status=status.HTTP_200_OK)
        except Cocktail.DoesNotExist:
            return Response(
                {"error": f"No se encontró un cóctel con id_cocktail={cocktail_id}"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

class OrdenesTodasView(APIView):
    def get(self, request):
        try:
            ordenes = Ordenes.objects.all()
            serializer = OrdenesSerializer(ordenes, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"Error al obtener todas las órdenes: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateIngredienteView(APIView):
    def post(self, request, ingrediente_id):
        try:
            ingrediente = Ingredientes.objects.get(id_ingredientes=ingrediente_id)
            serializer = IngredientesSerializer(ingrediente, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "message": "Ingrediente actualizado correctamente",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Ingredientes.DoesNotExist:
            return Response(
                {"error": f"No se encontró un ingrediente con ID {ingrediente_id}"},
                status=status.HTTP_404_NOT_FOUND
            )
        except IntegrityError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class IngredienteDetailView(APIView):
    def get(self, request, pk):
        try:
            ingrediente = Ingredientes.objects.get(pk=pk)
            serializer = IngredientesSerializer(ingrediente)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Ingredientes.DoesNotExist:
            return Response(
                {"error": f"No se encontró un ingrediente con ID {pk}"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request, pk):
        try:
            ingrediente = Ingredientes.objects.get(pk=pk)
            serializer = IngredientesSerializer(ingrediente, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Ingredientes.DoesNotExist:
            return Response(
                {"error": f"No se encontró un ingrediente con ID {pk}"},
                status=status.HTTP_404_NOT_FOUND
            )
        except IntegrityError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def patch(self, request, pk):
        try:
            ingrediente = Ingredientes.objects.get(pk=pk)
            serializer = IngredientesSerializer(ingrediente, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Ingredientes.DoesNotExist:
            return Response(
                {"error": f"No se encontró un ingrediente con ID {pk}"},
                status=status.HTTP_404_NOT_FOUND
            )
        except IntegrityError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CrearCocktailIngredientesView(APIView):
    def post(self, request):
        serializer = CrearCocktailIngredientesSerializer(data=request.data)
        if serializer.is_valid():
            try:
                nuevo_registro = serializer.save()
                return Response({
                    "mensaje": "Ingrediente agregado al cóctel exitosamente.",
                    "data": serializer.data
                }, status=status.HTTP_201_CREATED)
            except IntegrityError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({
                    "error": f"Error interno del servidor: {str(e)}"
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UpdateCocktailIngredienteView(APIView):
    def patch(self, request, pk):
        try:
            cocktail_ingrediente = CocktailIngredientes.objects.get(pk=pk)
            serializer = UpdateCocktailIngredienteSerializer(
                cocktail_ingrediente, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "message": "Ingrediente actualizado correctamente.",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except CocktailIngredientes.DoesNotExist:
            return Response({"error": "El registro no existe."}, status=status.HTTP_404_NOT_FOUND)
        except IntegrityError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CrearIngredienteView(APIView):
    def post(self, request):
        serializer = CrearIngredienteSerializer(data=request.data)
        if serializer.is_valid():
            try:
                ingrediente = serializer.save()
                return Response(
                    {"message": "Ingrediente creado exitosamente.", "data": serializer.data},
                    status=status.HTTP_201_CREATED,
                )
            except IntegrityError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({
                    "error": f"Error al crear el ingrediente: {str(e)}"
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CrearCocktailView(APIView):
    def post(self, request):
        serializer = CrearCocktailSerializer(data=request.data)
        if serializer.is_valid():
            try:
                cocktail = serializer.save()
                return Response(
                    {"message": "Cocktail creado exitosamente.", "data": serializer.data},
                    status=status.HTTP_201_CREATED,
                )
            except IntegrityError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({
                    "error": f"Error al crear el cocktail: {str(e)}"
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

from openpyxl.utils import get_column_letter

class GenerarConsumosHistoricosView(View):
    def get(self, request):
        try:
            # Fecha de corte: último día del mes completo anterior
            fecha_actual = datetime.date.today()
            fecha_fin = fecha_actual.replace(day=1) - datetime.timedelta(days=1)

            # Primer registro
            fecha_inicio = OrdenElementos.objects.aggregate(
                primera_fecha=Min("fechaorden")
            )["primera_fecha"]
            if not fecha_inicio:
                return JsonResponse({"error": "No hay registros en la base de datos."}, status=400)

            # 1) Acumular consumo total por (ingrediente, año, mes)
            consumo_total = {}

            # Consumo directo
            directos = (
                OrdenElementos.objects
                    .filter(escocktail=False, fechaorden__range=[fecha_inicio, fecha_fin])
                    .annotate(
                        año=F("fechaorden__year"),
                        mes=F("fechaorden__month"),
                        ingrediente=F("id_ingredientes__nombre")
                    )
                    .values("ingrediente", "año", "mes")
                    .annotate(total_litros=Sum(F("cantidad") * F("id_ingredientes__litrosporunidad")))
            )
            for d in directos:
                key = (d["ingrediente"], d["año"], d["mes"])
                consumo_total[key] = consumo_total.get(key, 0) + d["total_litros"]

            # Consumo por cócteles
            cocktails = OrdenElementos.objects.filter(
                escocktail=True, fechaorden__range=[fecha_inicio, fecha_fin]
            )
            for pedido in cocktails:
                receta = CocktailIngredientes.objects.filter(
                    id_cocktail=pedido.id_cocktail, activo=True
                )
                for ingr in receta:
                    key = (
                        ingr.id_ingredientes.nombre,
                        pedido.fechaorden.year,
                        pedido.fechaorden.month
                    )
                    consumo_total[key] = consumo_total.get(key, 0) + (pedido.cantidad * ingr.cantidad)

            # 2) Construir DataFrame “flat” con todas las combinaciones
            ingredientes = Ingredientes.objects.values_list("nombre", flat=True).distinct()
            fecha_fin_dt = pd.to_datetime(fecha_fin)
            años = range(fecha_inicio.year, fecha_fin_dt.year + 1)
            meses = range(1, 13)
            combinaciones = []
            for año in años:
                if año == fecha_fin_dt.year:
                    meses_hasta = range(1, fecha_fin_dt.month + 1)
                else:
                    meses_hasta = meses
                combinaciones.extend(product(ingredientes, [año], meses_hasta))

            datos = [
                {
                    "Ingrediente": ing,
                    "Año": año,
                    "Mes": mes,
                    "Consumo Total": float(consumo_total.get((ing, año, mes), 0))
                }
                for ing, año, mes in combinaciones
            ]
            df = pd.DataFrame(datos)
            df.sort_values(by=["Ingrediente", "Año", "Mes"], inplace=True)

            # 3) Pivot table
            pivot = (
                df
                .pivot(index=["Ingrediente", "Año"], columns="Mes", values="Consumo Total")
                .reindex(columns=range(1, 13), fill_value=0)
                .astype(float)
            )
            meses_abv = {
                1:"Ene",  2:"Feb", 3:"Mar", 4:"Abr",
                5:"May",  6:"Jun", 7:"Jul", 8:"Ago",
                9:"Sep", 10:"Oct",11:"Nov",12:"Dic"
            }
            pivot.rename(columns=meses_abv, inplace=True)

            # 4) Exportar a Excel con auto-width
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                ruta = tmp.name
                with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
                    pivot.to_excel(writer, sheet_name="Consumo Histórico")
                    ws = writer.sheets["Consumo Histórico"]
                    for col in ws.columns:
                        max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
                        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

            # 5) Devolver .xlsx
            archivo = open(ruta, "rb")
            return FileResponse(
                archivo,
                as_attachment=True,
                filename="consumo_historico.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

class SubirConsumoHistoricoView(View):
    def post(self, request):
        try:
            archivo = request.FILES.get('archivo')
            if not archivo:
                return JsonResponse({"error": "No se proporcionó ningún archivo."}, status=400)

            fs = FileSystemStorage(location=settings.MEDIA_ROOT)
            nombre_archivo = "consumo_historico.xlsx"
            ruta_archivo = fs.save(nombre_archivo, archivo)

            return JsonResponse({"mensaje": f"Archivo {ruta_archivo} subido exitosamente."}, status=200)
        except Exception as e:
            return JsonResponse({"error": f"Error al subir el archivo: {str(e)}"}, status=500)

from sklearn.metrics import mean_absolute_error, mean_squared_error

@method_decorator(csrf_exempt, name="dispatch")
class GenerarPrediccionesView(View):
    def post(self, request):
        try:
            # --- 1. EXTRAER Y PREPARAR DATOS HISTÓRICOS ---
            fecha_actual = datetime.date.today()
            fecha_fin = fecha_actual.replace(day=1) - datetime.timedelta(days=1)
            fecha_inicio = OrdenElementos.objects.aggregate(
                primera_fecha=Min("fechaorden")
            )["primera_fecha"]
            if not fecha_inicio:
                return JsonResponse({"error": "No hay registros en la base de datos."}, status=400)

            consumo_total = {}
            # Consumos directos
            for d in (
                OrdenElementos.objects
                    .filter(escocktail=False, fechaorden__range=[fecha_inicio, fecha_fin])
                    .annotate(año=F("fechaorden__year"),
                              mes=F("fechaorden__month"),
                              ingrediente=F("id_ingredientes__nombre"))
                    .values("ingrediente","año","mes")
                    .annotate(total_litros=Sum(F("cantidad")*F("id_ingredientes__litrosporunidad")))
            ):
                key = (d["ingrediente"], d["año"], d["mes"])
                consumo_total[key] = consumo_total.get(key,0) + float(d["total_litros"])

            # Consumos por cóctel
            for ped in OrdenElementos.objects.filter(
                escocktail=True, fechaorden__range=[fecha_inicio, fecha_fin]
            ):
                for rec in CocktailIngredientes.objects.filter(
                    id_cocktail=ped.id_cocktail, activo=True
                ):
                    key = (
                        rec.id_ingredientes.nombre,
                        ped.fechaorden.year,
                        ped.fechaorden.month
                    )
                    consumo_total[key] = consumo_total.get(key,0) + float(ped.cantidad*rec.cantidad)

            # Todas las combinaciones (ingrediente, año, mes)
            ingredientes = list(Ingredientes.objects.values_list("nombre",flat=True).distinct())
            fecha_fin_dt = pd.to_datetime(fecha_fin)
            años = range(fecha_inicio.year, fecha_fin_dt.year+1)
            meses = range(1,13)
            combos = []
            for año in años:
                hasta = range(1, fecha_fin_dt.month+1) if año==fecha_fin_dt.year else meses
                combos += product(ingredientes,[año],hasta)

            datos = [
                {"Ingrediente":ing,"Año":año,"Mes":mes,
                 "Consumo Total": consumo_total.get((ing,año,mes),0)}
                for ing,año,mes in combos
            ]
            df_hist = pd.DataFrame(datos).sort_values(["Ingrediente","Año","Mes"])
            df_hist["Consumo Total"].fillna(0, inplace=True)

            # Pivot table histórico
            pivot_hist = (
                df_hist
                .pivot(index=["Ingrediente","Año"],columns="Mes",values="Consumo Total")
                .reindex(columns=range(1,13),fill_value=0)
                .astype(float)
            )
            meses_abv = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",
                         5:"May",6:"Jun",7:"Jul",8:"Ago",
                         9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
            pivot_hist.rename(columns=meses_abv, inplace=True)

            # --- 2. MODELADO, PRONÓSTICO Y MÉTRICAS ---
            n_pron = 12
            lista_preds = []
            lista_mets  = []

            for ing in ingredientes:
                # Construir la serie histórica
                periodos = []
                y, m = fecha_inicio.year, fecha_inicio.month
                while (y<fecha_fin_dt.year) or (y==fecha_fin_dt.year and m<=fecha_fin_dt.month):
                    periodos.append((y,m))
                    if m==12:
                        y, m = y+1, 1
                    else:
                        m += 1

                vec = [consumo_total.get((ing,año,mes),0) for año,mes in periodos]
                idx = pd.date_range(
                    start=f"{periodos[0][0]}-{periodos[0][1]:02d}-01",
                    periods=len(vec), freq="MS"
                )
                serie = pd.Series(vec, index=idx)
                serie_log = np.log1p(serie)

                # Grid search (AIC + BIC)
                best_aic, best_bic = np.inf, np.inf
                best_order, best_season = None, None
                best_mod = None
                for p in range(3):
                    for d in range(2):
                        for q in range(3):
                            for P in range(2):
                                for D in range(2):
                                    for Q in range(2):
                                        try:
                                            m_ = SARIMAX(
                                                serie_log,
                                                order=(p,d,q),
                                                seasonal_order=(P,D,Q,12),
                                                enforce_stationarity=False,
                                                enforce_invertibility=False
                                            ).fit(disp=False, maxiter=300, method="lbfgs")
                                            if (m_.aic < best_aic) and (m_.bic < best_bic):
                                                best_aic, best_bic = m_.aic, m_.bic
                                                best_order = (p,d,q)
                                                best_season = (P,D,Q,12)
                                                best_mod = m_
                                        except:
                                            continue
                if not best_mod:
                    continue

                # Pronóstico de 12 meses arrancando en el mes actual
                fstart = pd.Timestamp(fecha_actual.year, fecha_actual.month, 1)
                fechas = pd.date_range(start=fstart, periods=n_pron, freq="MS")
                fc = best_mod.get_forecast(steps=n_pron)
                yhat = np.expm1(fc.predicted_mean)
                ci = np.expm1(fc.conf_int())

                dfp = pd.DataFrame({
                    "Ingrediente": ing,
                    "Fecha": fechas,
                    "Consumo Predicho": yhat.values,
                    "Límite Inferior": ci.iloc[:,0].values,
                    "Límite Superior": ci.iloc[:,1].values
                })
                lista_preds.append(dfp)

                # Validación cruzada (80/20 walk‐forward)
                s = serie_log
                n = len(s)
                ntr = int(n*0.8)
                tr, te = s.iloc[:ntr], s.iloc[ntr:]
                history, wp = list(tr), []
                for t in range(len(te)):
                    try:
                        cv = SARIMAX(
                            history,
                            order=best_order,
                            seasonal_order=best_season,
                            enforce_stationarity=False,
                            enforce_invertibility=False
                        ).fit(disp=False, maxiter=300, method="lbfgs")
                    except:
                        break
                    fcv = np.expm1(cv.get_forecast(1).predicted_mean)[0]
                    wp.append(fcv)
                    history.append(te.iloc[t])
                actuals = np.expm1(te.values)
                lista_mets.append({
                    "Ingrediente": ing,
                    "MAE (L)": mean_absolute_error(actuals, wp),
                    "RMSE (L)": np.sqrt(mean_squared_error(actuals, wp)),
                    "MAPE (%)": np.mean(np.abs((actuals - np.array(wp)) / actuals)) * 100
                })

            # DataFrames finales
            df_pred = pd.concat(lista_preds, ignore_index=True)
            df_pred["Año"] = df_pred["Fecha"].dt.year
            df_pred["Mes"] = df_pred["Fecha"].dt.month
            pivot_pred = (
                df_pred
                .pivot(index=["Ingrediente","Año"], columns="Mes", values="Consumo Predicho")
                .reindex(columns=range(1,13), fill_value=0)
                .astype(float)
                .rename(columns=meses_abv)
                .round(2)
            )
            df_mets = pd.DataFrame(lista_mets)

            # --- 3. HOJA REABASTECIMIENTO (próximos 3 meses) ---
            df_next3 = (
                df_pred
                .sort_values(["Ingrediente","Fecha"])
                .groupby("Ingrediente")
                .head(3)
                .copy()
            )
            stock_df = pd.DataFrame(
                list(Ingredientes.objects.values("nombre","cantidadactual"))
            ).rename(columns={"nombre":"Ingrediente","cantidadactual":"Stock Actual"})
            # convertir stock a float para no mezclar Decimal/float
            stock_map = {row["Ingrediente"]: float(row["Stock Actual"]) for _, row in stock_df.iterrows()}

            restock = []
            for ing, grp in df_next3.groupby("Ingrediente"):
                rem = stock_map.get(ing, 0.0)
                for _, row in grp.iterrows():
                    need = max(row["Consumo Predicho"] - max(rem, 0.0), 0.0)
                    restock.append({
                        "Ingrediente": ing,
                        "Fecha": row["Fecha"],
                        "Stock Inicial": stock_map.get(ing, 0.0),
                        "Consumo Predicho": row["Consumo Predicho"],
                        "Restock Necesario": need
                    })
                    rem -= row["Consumo Predicho"]
            df_reab = pd.DataFrame(restock)

            # --- 4. EXPORTAR TODO A UN .XLSX con 4 hojas ---
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                path = tmp.name
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    pivot_hist.to_excel(writer, sheet_name="Consumo Histórico")
                    pivot_pred.to_excel(writer, sheet_name="Predicciones")
                    df_mets.to_excel(writer, index=False, sheet_name="Métricas")
                    df_reab.to_excel(writer, index=False, sheet_name="Reabastecimiento")
                    # Auto-width en cada columna
                    for ws in writer.sheets.values():
                        for col in ws.columns:
                            mx = max(len(str(c.value)) for c in col if c.value is not None)
                            ws.column_dimensions[get_column_letter(col[0].column)].width = mx + 2

            f = open(path, "rb")
            return FileResponse(
                f,
                as_attachment=True,
                filename="consumos_predicciones_metricas_reabastecimiento.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

class GenerarStockActualView(View):
    def get(self, request):
        try:
            ingredientes_stock = Ingredientes.objects.values("nombre", "cantidadactual")
            if not ingredientes_stock:
                return JsonResponse({"error": "No hay datos de stock actual disponibles."}, status=400)

            df = pd.DataFrame(list(ingredientes_stock))
            df.rename(columns={"nombre": "Ingrediente", "cantidadactual": "Cantidad Actual"}, inplace=True)

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                ruta_archivo = tmp_file.name
                df.to_excel(ruta_archivo, index=False, engine='openpyxl')

            archivo = open(ruta_archivo, "rb")
            response = FileResponse(archivo, as_attachment=True, filename="stock_actual.xlsx")
            return response

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@method_decorator(csrf_exempt, name="dispatch")
class GenerarReporteReabastecimientoView(View):
    def post(self, request):
        try:
            fecha_actual = datetime.date.today()
            fecha_fin = fecha_actual.replace(day=1) - datetime.timedelta(days=1)

            fecha_inicio = OrdenElementos.objects.aggregate(primera_fecha=Min("fechaorden"))["primera_fecha"]

            if not fecha_inicio:
                return JsonResponse({"error": "No hay registros en la base de datos."}, status=400)

            consumo_total = {}

            datos_directos = (
                OrdenElementos.objects
                .filter(escocktail=False, fechaorden__range=[fecha_inicio, fecha_fin])
                .annotate(
                    año=F('fechaorden__year'),
                    mes=F('fechaorden__month'),
                    ingrediente_nombre=F('id_ingredientes__nombre')
                )
                .values('ingrediente_nombre', 'año', 'mes')
                .annotate(total_litros=Sum(F('cantidad') * F('id_ingredientes__litrosporunidad')))
            )

            for dato in datos_directos:
                key = (dato['ingrediente_nombre'], dato['año'], dato['mes'])
                consumo_total[key] = consumo_total.get(key, 0) + dato['total_litros']

            datos_cocktails = OrdenElementos.objects.filter(
                escocktail=True,
                fechaorden__range=[fecha_inicio, fecha_fin]
            )

            for cocktail in datos_cocktails:
                receta = CocktailIngredientes.objects.filter(
                    id_cocktail=cocktail.id_cocktail,
                    activo=True
                )
                for ingrediente in receta:
                    key = (ingrediente.id_ingredientes.nombre, cocktail.fechaorden.year, cocktail.fechaorden.month)
                    cantidad_usada = cocktail.cantidad * ingrediente.cantidad
                    consumo_total[key] = consumo_total.get(key, 0) + cantidad_usada

            ingredientes = OrdenElementos.objects.values_list("id_ingredientes__nombre", flat=True).distinct()
            fecha_fin_datetime = pd.to_datetime(fecha_fin)
            años = range(fecha_inicio.year, fecha_fin_datetime.year + 1)
            meses = range(1, 13)

            combinaciones = []
            for año in años:
                if año == fecha_fin_datetime.year:
                    meses_hasta_fin = range(1, fecha_fin_datetime.month + 1)
                    combinaciones.extend(product(ingredientes, [año], meses_hasta_fin))
                else:
                    combinaciones.extend(product(ingredientes, [año], meses))

            datos_consumo = [
                {
                    "Ingrediente": ingrediente,
                    "Año": año,
                    "Mes": mes,
                    "Consumo Total": float(consumo_total.get((ingrediente, año, mes), 0))
                }
                for ingrediente, año, mes in combinaciones
            ]

            df_consumo = pd.DataFrame(datos_consumo)
            df_consumo.sort_values(by=['Ingrediente', 'Año', 'Mes'], inplace=True)

            df_consumo["Consumo Total"] = df_consumo["Consumo Total"].fillna(0)

            predicciones_totales = pd.DataFrame(
                columns=["Ingrediente", "Fecha", "Consumo Predicho", "Límite Inferior", "Límite Superior"]
            )

            for ingrediente in df_consumo["Ingrediente"].unique():
                datos_ingrediente = df_consumo[df_consumo["Ingrediente"] == ingrediente]
                if datos_ingrediente.empty or datos_ingrediente["Consumo Total"].sum() == 0:
                    continue

                serie = pd.Series(
                    data=datos_ingrediente["Consumo Total"].values,
                    index=pd.date_range(
                        start=f"{int(datos_ingrediente['Año'].min())}-{int(datos_ingrediente['Mes'].min()):02d}-01",
                        periods=len(datos_ingrediente),
                        freq="M",
                    ),
                )
                serie_log = np.log1p(serie)
                modelo = SARIMAX(serie_log, order=(1, 1, 1), seasonal_order=(1, 1, 1, 12))
                resultado = modelo.fit(disp=False)

                prediccion = resultado.get_forecast(steps=12)
                prediccion_medias = np.expm1(prediccion.predicted_mean)
                intervalo_confianza = np.expm1(prediccion.conf_int())

                df_predicciones = pd.DataFrame({
                    "Ingrediente": ingrediente,
                    "Fecha": pd.date_range(start=serie.index[-1] + pd.DateOffset(months=1), periods=12, freq="M"),
                    "Consumo Predicho": prediccion_medias.values,
                    "Límite Inferior": intervalo_confianza.iloc[:, 0].values,
                    "Límite Superior": intervalo_confianza.iloc[:, 1].values,
                })
                predicciones_totales = pd.concat([predicciones_totales, df_predicciones], ignore_index=True)

            predicciones_totales["Fecha"] = pd.to_datetime(predicciones_totales["Fecha"])
            predicciones_totales = predicciones_totales.sort_values(by=["Ingrediente", "Fecha"])
            predicciones_filtradas = predicciones_totales.groupby("Ingrediente").head(3)

            stock_actual = Ingredientes.objects.values("nombre", "cantidadactual")
            stock_actual_df = pd.DataFrame(stock_actual)
            stock_actual_df.rename(columns={"nombre": "Ingrediente", "cantidadactual": "Cantidad Actual"}, inplace=True)

            reporte = predicciones_filtradas[["Ingrediente", "Fecha", "Consumo Predicho"]].copy()
            reporte["Cantidad Actual"] = reporte["Ingrediente"].map(
                dict(zip(stock_actual_df["Ingrediente"], stock_actual_df["Cantidad Actual"]))
            ).fillna(0)

            reporte["Rebastecimiento Recomendado"] = (
                reporte["Consumo Predicho"] * 1.1 - reporte["Cantidad Actual"].apply(float)
            ).clip(lower=0).round()

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                ruta_reporte = tmp_file.name
                with pd.ExcelWriter(ruta_reporte, engine="openpyxl") as writer:
                    df_consumo.to_excel(writer, index=False, sheet_name="Consumo Histórico")
                    predicciones_totales.to_excel(writer, index=False, sheet_name="Predicciones Anuales")
                    reporte.to_excel(writer, index=False, sheet_name="Reabastecimiento")

            archivo = open(ruta_reporte, "rb")
            return FileResponse(
                archivo,
                as_attachment=True,
                filename="reporte_consumos_reabastecimiento.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


class GenerarReporteEstadisticoExcelView(APIView):
    def post(self, request):
        try:
            primer_fecha = Ordenes.objects.earliest("fechahora").fechahora
            fecha_actual = datetime.date.today()
            fecha_fin = fecha_actual.replace(day=1) - datetime.timedelta(days=1)

            datos_reporte = {}
            for año in range(primer_fecha.year, fecha_fin.year + 1):
                datos_anuales = []
                for mes in range(1, 13):
                    if año == fecha_fin.year and mes > fecha_fin.month:
                        break

                    fecha_inicio = datetime.date(año, mes, 1)
                    fecha_termino = (
                        datetime.date(año, mes + 1, 1)
                        if mes < 12
                        else datetime.date(año + 1, 1, 1)
                    )

                    bebida_favorita = OrdenElementos.objects.filter(
                        id_orden__fechahora__range=[fecha_inicio, fecha_termino]
                    ).annotate(
                        nombreBebida=Case(
                            When(escocktail=True, then="id_cocktail__nombre"),
                            When(escocktail=False, then="id_ingredientes__nombre"),
                            output_field=CharField(),
                        )
                    ).values("nombreBebida").annotate(
                        vecesOrdenado=Count("id_elemento")
                    ).order_by("-vecesOrdenado").first()

                    bar_mas_consumo = Ordenes.objects.filter(
                        fechahora__range=[fecha_inicio, fecha_termino]
                    ).values("id_bar__nombre").annotate(
                        totalOrdenes=Count("id_bar")
                    ).order_by("-totalOrdenes").first()

                    nacionalidad_predominante = AsignacionesHuespedes.objects.filter(
                        fechaasignacion__range=[fecha_inicio, fecha_termino]
                    ).values("id_huesped__nacionalidad").annotate(
                        totalReservas=Count("id_asignacion")
                    ).order_by("-totalReservas").first()

                    ingrediente_mas_ordenado = OrdenElementos.objects.filter(
                        escocktail=False,
                        id_orden__fechahora__range=[fecha_inicio, fecha_termino]
                    ).values("id_ingredientes__nombre").annotate(
                        vecesOrdenado=Count("id_elemento")
                    ).order_by("-vecesOrdenado").first()

                    coctel_mas_ordenado = OrdenElementos.objects.filter(
                        escocktail=True,
                        id_orden__fechahora__range=[fecha_inicio, fecha_termino]
                    ).values("id_cocktail__nombre").annotate(
                        vecesOrdenado=Count("id_elemento")
                    ).order_by("-vecesOrdenado").first()

                    consumo_ingredientes = OrdenElementos.objects.filter(
                        escocktail=False,
                        id_orden__fechahora__range=[fecha_inicio, fecha_termino],
                    ).aggregate(total=Sum("cantidad"))["total"] or 0

                    consumo_cocteles = OrdenElementos.objects.filter(
                        escocktail=True,
                        id_orden__fechahora__range=[fecha_inicio, fecha_termino],
                    ).aggregate(total=Sum("cantidad"))["total"] or 0

                    datos_anuales.append({
                        "Mes": f"{mes}/{año}",
                        "Bebida Favorita": bebida_favorita["nombreBebida"] if bebida_favorita else "N/A",
                        "Veces Ordenada": bebida_favorita["vecesOrdenado"] if bebida_favorita else 0,
                        "Bar Más Consumo": bar_mas_consumo["id_bar__nombre"] if bar_mas_consumo else "N/A",
                        "Nacionalidad Predominante": nacionalidad_predominante["id_huesped__nacionalidad"] if nacionalidad_predominante else "N/A",
                        "Ingrediente Más Ordenado": ingrediente_mas_ordenado["id_ingredientes__nombre"] if ingrediente_mas_ordenado else "N/A",
                        "Cóctel Más Ordenado": coctel_mas_ordenado["id_cocktail__nombre"] if coctel_mas_ordenado else "N/A",
                        "Total Ingredientes Consumidos": consumo_ingredientes,
                        "Total Cócteles Consumidos": consumo_cocteles,
                    })

                datos_reporte[año] = datos_anuales

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                with pd.ExcelWriter(tmp_file.name, engine="openpyxl") as writer:
                    for año, datos in datos_reporte.items():
                        df = pd.DataFrame(datos)
                        df.to_excel(writer, index=False, sheet_name=str(año))

                archivo = open(tmp_file.name, "rb")
                return FileResponse(
                    archivo,
                    as_attachment=True,
                    filename="reporte_estadistico.xlsx",
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
        

class HuespedesListView(APIView):
    def get(self, request):
        try:
            pago_realizado_subquery = AsignacionesHuespedes.objects.filter(
                id_huesped=OuterRef("id_huesped"),
                enhotel=True
            ).values('pagorealizado')[:1]

            huespedes = Huesped.objects.annotate(
                en_hotel=Exists(
                    AsignacionesHuespedes.objects.filter(
                        id_huesped=OuterRef("id_huesped"),
                        enhotel=True
                    )
                ),
                pagorealizado=Subquery(pago_realizado_subquery)
            )
            serializer = HuespedSerializer(huespedes, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CrearHuespedView(APIView):
    def post(self, request):
        serializer = HuespedSerializer(data=request.data)
        if serializer.is_valid():
            try:
                huesped = serializer.save()
                return Response({
                    "mensaje": "Huésped creado exitosamente.",
                    "id_huesped": huesped.id_huesped
                }, status=status.HTTP_201_CREATED)
            except IntegrityError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CrearHabitacionView(APIView):
    def post(self, request):
        serializer = HabitacionesSerializer(data=request.data)
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(
                    {"message": "Habitación creada exitosamente.", "data": serializer.data},
                    status=status.HTTP_201_CREATED,
                )
            except IntegrityError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response(
                    {"error": f"Error al crear la habitación: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class HabitacionesDisponiblesView(APIView):
    def get(self, request):
        try:
            habitaciones = Habitaciones.objects.all()
            serializer = HabitacionDisponibleSerializer(habitaciones, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"Error al obtener las habitaciones disponibles: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        

class AsignacionesHuespedesListView(APIView):
    def get(self, request):
        try:
            asignaciones = AsignacionesHuespedes.objects.all()
            serializer = AsignacionesHuespedesDetailSerializer(asignaciones, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"Error al obtener las asignaciones de huéspedes: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AsignacionActivaPorHuespedView(APIView):
    def get(self, request, id_huesped):
        try:
            asignacion = AsignacionesHuespedes.objects.filter(id_huesped=id_huesped, enhotel=True).first()
            if asignacion:
                serializer = AsignacionesHuespedesSerializer(asignacion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response({"mensaje": "No hay asignación activa para este huésped."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HuespedDetailView(APIView):
    def get(self, request, id_huesped):
        try:
            huesped = Huesped.objects.get(id_huesped=id_huesped)
            serializer = HuespedSerializer(huesped)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Huesped.DoesNotExist:
            return Response({"error": "Huésped no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        except IntegrityError as e:
            return Response({"error": f"Error de integridad de la base de datos: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CrearAsignacionView(APIView):
    def post(self, request):
        serializer = CrearAsignacionSerializer(data=request.data)
        if serializer.is_valid():
            try:
                asignacion = serializer.save()
                return Response({
                    "mensaje": "Asignación registrada exitosamente",
                    "ID_Asignacion": asignacion.id_asignacion
                }, status=status.HTTP_201_CREATED)
            except IntegrityError as e:
                return Response({
                    "error": f"Error de integridad al guardar la asignación: {str(e)}"
                }, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({
                    "error": f"Error al registrar la asignación: {str(e)}"
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ActualizarPagoAsignacionView(APIView):
    def patch(self, request, id_asignacion):
        try:
            asignacion = AsignacionesHuespedes.objects.get(id_asignacion=id_asignacion)
            asignacion.pagorealizado = request.data["pago_realizado"]
            asignacion.save()
            return Response({"mensaje": "Estado de pago actualizado correctamente"}, status=status.HTTP_200_OK)
        except AsignacionesHuespedes.DoesNotExist:
            return Response({"error": "Asignación no encontrada"}, status=status.HTTP_404_NOT_FOUND)
        except IntegrityError as e:
            return Response({"error": f"Error de integridad de la base de datos: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CheckoutHuespedView(APIView):
    def patch(self, request, id_asignacion):
        try:
            asignacion = AsignacionesHuespedes.objects.get(id_asignacion=id_asignacion)
            asignacion.enhotel = False
            asignacion.fechacheckout = now()
            asignacion.save()
            return Response({"mensaje": "Checkout realizado exitosamente"}, status=status.HTTP_200_OK)
        except AsignacionesHuespedes.DoesNotExist:
            return Response({"error": "Asignación no encontrada"}, status=status.HTTP_404_NOT_FOUND)
        except IntegrityError as e:
            return Response({"error": f"Error de integridad de la base de datos: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CambiarActividadOrdenView(APIView):
    def patch(self, request, pk):
        try:
            orden = Ordenes.objects.get(pk=pk)
            actividad = request.data.get('actividad')
            if actividad is not None:
                orden.actividad = actividad
                orden.save()
                return Response({"mensaje": "Actividad de la orden actualizada correctamente"}, status=status.HTTP_200_OK)
            else:
                return Response({"error": "No se proporcionó el estado de actividad"}, status=status.HTTP_400_BAD_REQUEST)
        except Ordenes.DoesNotExist:
            return Response({"error": "Orden no encontrada"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class OrdenesPorAsignacionView(APIView):
    def get(self, request, id_asignacion):
        try:
            ordenes = Ordenes.objects.filter(id_asignacion=id_asignacion, actividad=True)
            serializer = OrdenesSerializer(ordenes, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmpleadoListCreateAPIView(APIView):
    def get(self, request):
        try:
            empleados = Empleados.objects.all()
            serializer = EmpleadoSerializer(empleados, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        serializer = EmpleadoSerializer(data=request.data)
        if serializer.is_valid():
            try:
                empleado = serializer.save()
                return Response(
                    {"mensaje": "Empleado creado exitosamente", "data": serializer.data},
                    status=status.HTTP_201_CREATED
                )
            except IntegrityError as e:
                return Response({"error": f"Error de integridad: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BarCreateAPIView(APIView):
    def post(self, request):
        serializer = BaresSerializer(data=request.data)
        if serializer.is_valid():
            try:
                bar = serializer.save()
                return Response(
                    {"mensaje": "Bar creado exitosamente", "data": serializer.data},
                    status=status.HTTP_201_CREATED
                )
            except IntegrityError as e:
                return Response({"error": f"Error de integridad: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RegistrarMovimientoStockView(APIView):
    def post(self, request, ingrediente_id):
        try:
            serializer = MovimientoStockSerializer(data=request.data)
            if serializer.is_valid():
                tipo_movimiento = serializer.validated_data['tipomovimiento']
                cantidad = serializer.validated_data['cantidad']

                ingrediente = Ingredientes.objects.filter(id_ingredientes=ingrediente_id, activo=True).first()
                if not ingrediente:
                    return Response(
                        {"error": "Ingrediente no encontrado o está inactivo."},
                        status=status.HTTP_404_NOT_FOUND
                    )
                movimiento = MovimientosStock.objects.create(
                    id_ingredientes=ingrediente,
                    tipomovimiento=tipo_movimiento,
                    cantidad=cantidad,
                )
                return Response(
                    {
                        "mensaje": "Movimiento registrado exitosamente.",
                        "data": MovimientoStockSerializer(movimiento).data,
                    },
                    status=status.HTTP_201_CREATED
                )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except IntegrityError as e:
            return Response(
                {"error": f"Error de integridad: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            print("Error interno del servidor:", str(e))  
            return Response(
                {"error": f"Error interno del servidor: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ActualizarTandaOrdenView(APIView):
    def post(self, request, id_orden):
        try:
            num_cocktails = request.data.get("num_cocktails", 0)
            num_ingredientes = request.data.get("num_ingredientes", 0)

            orden = Ordenes.objects.get(id_orden=id_orden)
            
            orden.num_cocktails = num_cocktails
            orden.num_ingredientes = num_ingredientes
            orden.save()

            return Response({
                "mensaje": "Orden actualizada correctamente.",
                "id_orden": orden.id_orden,
                "num_cocktails": orden.num_cocktails,
                "num_ingredientes": orden.num_ingredientes,
            }, status=status.HTTP_200_OK)

        except Ordenes.DoesNotExist:
            return Response({
                "error": f"No se encontró la orden con id_orden={id_orden}"
            }, status=status.HTTP_404_NOT_FOUND)
        except ValidationError as ve:
            return Response({"error": ve.messages}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "error": f"Error al actualizar la orden: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CerrarOrdenesPorAsignacionView(APIView):
    def patch(self, request, id_asignacion):
        try:
            asignacion = AsignacionesHuespedes.objects.select_related('numerohabitacion').get(id_asignacion=id_asignacion)

            habitacion = asignacion.numerohabitacion  
            ordenes_activas = Ordenes.objects.filter(id_asignacion=asignacion.id_asignacion, actividad=True)

            if not ordenes_activas.exists():
                return Response(
                    {"mensaje": "No hay órdenes activas para esta asignación."},
                    status=status.HTTP_200_OK
                )
            ordenes_activas.update(actividad=False)
            ordenes_cerradas = Ordenes.objects.filter(id_asignacion=asignacion.id_asignacion, actividad=False)
            serializer_ordenes = OrdenesSerializer(ordenes_cerradas, many=True)
            return Response(
                {
                    "mensaje": "Todas las órdenes asociadas a la asignación han sido cerradas correctamente.",
                    "ordenes_cerradas": serializer_ordenes.data,
                },
                status=status.HTTP_200_OK
            )
        except AsignacionesHuespedes.DoesNotExist:
            return Response({"error": "Asignación no encontrada."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": f"Error al cerrar las órdenes: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateEmpleadoActivoView(APIView):
    def patch(self, request, pk, format=None):
        try:
            empleado = Empleados.objects.get(pk=pk)
        except Empleados.DoesNotExist:
            return Response(
                {"error": f"No se encontró un empleado con ID {pk}."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        activo = request.data.get('activo', None)
        if activo is None:
            return Response(
                {"error": "El campo 'activo' es requerido."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = {'activo': activo}
        
        serializer = EmpleadoSerializer(empleado, data=data, partial=True)
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(
                    {
                        "mensaje": "Estado activo del empleado actualizado correctamente.",
                        "data": serializer.data
                    },
                    status=status.HTTP_200_OK
                )
            except IntegrityError as e:
                return Response(
                    {"error": f"Error de integridad: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            except Exception as e:
                return Response(
                    {"error": f"Error al actualizar el empleado: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UpdateBarActivoView(APIView):
    def patch(self, request, pk, format=None):
        try:
            bar = Bares.objects.get(pk=pk)
        except Bares.DoesNotExist:
            return Response(
                {"error": f"No se encontró un bar con ID {pk}."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        activo = request.data.get('activo', None)
        if activo is None:
            return Response(
                {"error": "El campo 'activo' es requerido."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = {'activo': activo}
        
        serializer = BaresSerializer(bar, data=data, partial=True)
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(
                    {
                        "mensaje": "Estado activo del bar actualizado correctamente.",
                        "data": serializer.data
                    },
                    status=status.HTTP_200_OK
                )
            except IntegrityError as e:
                return Response(
                    {"error": f"Error de integridad: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            except Exception as e:
                return Response(
                    {"error": f"Error al actualizar el bar: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UpdateHuespedActivoView(APIView):
    def patch(self, request, pk, format=None):
        try:
            huesped = Huesped.objects.get(pk=pk)
        except Huesped.DoesNotExist:
            return Response(
                {"error": f"No se encontró un huésped con ID {pk}."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        activo = request.data.get('activo', None)
        if activo is None:
            return Response(
                {"error": "El campo 'activo' es requerido."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = {'activo': activo}
        
        serializer = HuespedSerializer(huesped, data=data, partial=True)
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(
                    {
                        "mensaje": "Estado activo del huésped actualizado correctamente.",
                        "data": serializer.data
                    },
                    status=status.HTTP_200_OK
                )
            except IntegrityError as e:
                return Response(
                    {"error": f"Error de integridad: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            except Exception as e:
                return Response(
                    {"error": f"Error al actualizar el huésped: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class RegistroOrdenesListView(APIView):
    def get(self, request):
        try:
            registroOrdenes = RegistroOrdenes.objects.all()
            serializer = RegistroOrdenesSerializer(registroOrdenes, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RegistroMovimientosListView(APIView):
    def get(self, request):
        try:
            registroMovimientos = RegistroMovimientos.objects.all()
            serializer = RegistroMovimientosSerializer(registroMovimientos, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class RegistroEstanciasListView(APIView):
    def get(self, request):
        try:
            registroEstancias = RegistroEstancias.objects.all()
            serializer = RegistroEstanciasSerializer(registroEstancias, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

