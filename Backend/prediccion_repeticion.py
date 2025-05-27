import os
import django
import pandas as pd
from django.db.models import Sum, F, Min, Max
import warnings
from itertools import product  
from openpyxl.utils import get_column_letter

# Filtrar warnings generales y de convergencia de statsmodels
from statsmodels.tools.sm_exceptions import ConvergenceWarning
warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", category=ConvergenceWarning)

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "hotel_management.settings")
django.setup()

from gestion.models import OrdenElementos, CocktailIngredientes, Ingredientes

def generar_consumos_historicos_excel():
    # Obtener fecha de inicio (primer registro) y de finalización (último registro)
    fecha_inicio = OrdenElementos.objects.aggregate(primera_fecha=Min("fechaorden"))["primera_fecha"]
    fecha_fin = OrdenElementos.objects.aggregate(ultima_fecha=Max("fechaorden"))["ultima_fecha"]

    if not fecha_inicio or not fecha_fin:
        print("No hay registros en la base de datos para calcular consumos.")
        return

    print(f"Fecha de inicio detectada automáticamente: {fecha_inicio}")
    print(f"Fecha de finalización detectada automáticamente: {fecha_fin}")

    consumo_total = {}

    # Cálculo de consumo para registros directos (no cocktail)
    datos_directos = (
        OrdenElementos.objects
        .filter(escocktail=False, fechaorden__range=[fecha_inicio, fecha_fin])
        .annotate(
            año=F('fechaorden__year'),
            mes=F('fechaorden__month'),
            ingrediente_nombre=F('id_ingredientes__nombre')
        )
        .values('ingrediente_nombre', 'año', 'mes')
        .annotate(total_litros=Sum(F('cantidad') * F('id_ingredientes__litrosporunidad')))
    )

    for dato in datos_directos:
        key = (dato['ingrediente_nombre'], dato['año'], dato['mes'])
        consumo_total[key] = consumo_total.get(key, 0) + dato['total_litros']

    # Cálculo de consumo para registros de cocktails
    datos_cocktails = OrdenElementos.objects.filter(
        escocktail=True,
        fechaorden__range=[fecha_inicio, fecha_fin]
    )
    for cocktail in datos_cocktails:
        receta = CocktailIngredientes.objects.filter(
            id_cocktail=cocktail.id_cocktail,
            activo=True
        )
        for ingrediente in receta:
            key = (ingrediente.id_ingredientes.nombre, cocktail.fechaorden.year, cocktail.fechaorden.month)
            cantidad_usada = cocktail.cantidad * ingrediente.cantidad
            consumo_total[key] = consumo_total.get(key, 0) + cantidad_usada

    # Obtener todos los ingredientes existentes en la base de datos
    ingredientes = Ingredientes.objects.values_list("nombre", flat=True).distinct()
    fecha_fin_datetime = pd.to_datetime(fecha_fin)
    años = range(fecha_inicio.year, fecha_fin_datetime.year + 1)
    meses = range(1, 13)

    combinaciones = []
    for año in años:
        if año == fecha_fin_datetime.year:
            meses_hasta_fin = range(1, fecha_fin_datetime.month + 1)
            combinaciones.extend(product(ingredientes, [año], meses_hasta_fin))
        else:
            combinaciones.extend(product(ingredientes, [año], meses))

    datos_consumo = [
        {
            "Ingrediente": ingrediente,
            "Año": año,
            "Mes": mes,
            "Consumo Total": consumo_total.get((ingrediente, año, mes), 0)
        }
        for ingrediente, año, mes in combinaciones
    ]

    df = pd.DataFrame(datos_consumo)
    df.sort_values(by=['Ingrediente', 'Año', 'Mes'], inplace=True)

    # Crear la tabla pivot: filas = (Ingrediente, Año) y columnas = Mes (1 a 12)
    pivot_table = df.pivot(index=["Ingrediente", "Año"], columns="Mes", values="Consumo Total")
    pivot_table = pivot_table.reindex(columns=range(1, 13), fill_value=0)
    pivot_table = pivot_table.astype(float)

    # Renombrar columnas a abreviaturas de 3 letras
    meses_abreviados = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
        5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
    }
    pivot_table.rename(columns=meses_abreviados, inplace=True)

    # --- Generación de vectores de consumos históricos ---
    periodos = []
    current_year = fecha_inicio.year
    current_month = fecha_inicio.month
    while current_year < fecha_fin.year or (current_year == fecha_fin.year and current_month <= fecha_fin.month):
        periodos.append((current_year, current_month))
        if current_month == 12:
            current_year += 1
            current_month = 1
        else:
            current_month += 1

    vectores_consumo = {}
    for ing in ingredientes:
        vector_consumo = [float(consumo_total.get((ing, año, mes), 0)) for (año, mes) in periodos]
        vectores_consumo[ing] = vector_consumo
        print(f"{ing}: {vector_consumo}")    

    # Exportar la tabla pivot a Excel con autoajuste de columnas
    ruta_archivo = "consumo_historico.xlsx"
    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        pivot_table.to_excel(writer, sheet_name="Consumos", index=True)
        worksheet = writer.sheets["Consumos"]
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    print(f"Archivo generado: {ruta_archivo}")
    return fecha_inicio, fecha_fin, vectores_consumo

# Ejecutar la función histórica y obtener vectores
fecha_inicio, fecha_fin, vectores_consumo = generar_consumos_historicos_excel()

# ========================================================================================
import numpy as np
from statsmodels.tsa.statespace.sarimax import SARIMAX
import matplotlib.pyplot as plt
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from sklearn.metrics import mean_absolute_error, mean_squared_error, mean_absolute_percentage_error

n_pronosticos = 12
predicciones_list = []
model_params = {}  # Para guardar los parámetros del mejor modelo de cada ingrediente
meses_abreviados = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
}

# Función para hacer validación cruzada tipo walk-forward (usando un 70% de la serie como entrenamiento inicial)
def walk_forward_validation(serie, order, seasonal_order, forecast_horizon=1, initial_train_ratio=0.7):
    actuals = []
    predictions = []
    n = len(serie)
    train_size = int(n * initial_train_ratio)
    for i in range(train_size, n - forecast_horizon + 1):
        train = serie.iloc[:i]
        test = serie.iloc[i:i+forecast_horizon]
        try:
            mod = SARIMAX(np.log1p(train), order=order, seasonal_order=seasonal_order,
                          enforce_stationarity=False, enforce_invertibility=False)
            res = mod.fit(disp=False, maxiter=300, method='lbfgs')
            forecast_log = res.get_forecast(steps=forecast_horizon).predicted_mean
            forecast = np.expm1(forecast_log)
            predictions.append(forecast.iloc[-1])
            actuals.append(test.iloc[-1])
        except Exception:
            continue
    return np.array(actuals), np.array(predictions)

# Bucle para cada ingrediente
for ing, vector in vectores_consumo.items():
    # Crear índice de fechas para la serie histórica
    index_historico = pd.date_range(start=fecha_inicio, periods=len(vector), freq='MS')
    serie = pd.Series(data=vector, index=index_historico)
    
    print(f"\n--- Serie histórica para {ing} ---")
    print(serie.head())
    print("...")
    print(serie.tail())
    
    # Si la serie tiene al menos 12 observaciones, se repiten los últimos 12 meses
    if len(serie) >= 12:
        ultimo_original = serie.index[-1]
        ultimos_12 = serie.iloc[-12:]
        nuevo_indice = pd.date_range(start=ultimo_original + pd.DateOffset(months=1), periods=12, freq='MS')
        bloque_repetido = pd.Series(data=ultimos_12.values, index=nuevo_indice)
        serie_modificada = pd.concat([serie, bloque_repetido]).sort_index()
        print(f"Para {ing}, se han repetido los últimos 12 meses en la serie.")
    else:
        serie_modificada = serie

    # Transformación logarítmica para estabilizar la varianza
    serie_log = np.log1p(serie_modificada)
    
    # ---------------------------
    # Fase de comparación de modelos usando AIC y BIC (Grid Search)
    # ---------------------------
    best_aic = np.inf
    best_bic = np.inf
    best_order = None
    best_seasonal_order = None
    best_model = None

    for p in range(0, 3):
        for d in range(0, 2):
            for q in range(0, 3):
                for P in range(0, 2):
                    for D in range(0, 2):
                        for Q in range(0, 2):
                            try:
                                mod = SARIMAX(serie_log, order=(p, d, q), seasonal_order=(P, D, Q, 12),
                                              enforce_stationarity=False, enforce_invertibility=False)
                                res = mod.fit(disp=False, maxiter=300, method='lbfgs')
                                if (res.aic < best_aic) and (res.bic < best_bic):
                                    best_aic = res.aic
                                    best_bic = res.bic
                                    best_order = (p, d, q)
                                    best_seasonal_order = (P, D, Q, 12)
                                    best_model = res
                            except Exception:
                                continue

    if best_model is None:
        print(f"No se pudo ajustar un modelo para {ing}.")
        continue

    print(f"Mejor modelo para {ing}: order={best_order}, seasonal_order={best_seasonal_order}, AIC={best_aic:.2f}, BIC={best_bic:.2f}")
    model_params[ing] = (best_order, best_seasonal_order)  # Guardar los parámetros

    # ---------------------------
    # Fase de visualización gráfica (diagnóstico)
    # ---------------------------
    forecast_start = fecha_fin.replace(day=1) + pd.DateOffset(months=1)
    forecast_index = pd.date_range(start=forecast_start, periods=n_pronosticos, freq='MS')
    forecast_obj = best_model.get_forecast(steps=n_pronosticos)
    prediccion_log = forecast_obj.predicted_mean
    prediccion = np.expm1(prediccion_log)

    # Gráficos diagnósticos: serie, pronóstico, ACF y PACF de residuos
    residuals = best_model.resid
    max_lags = min(20, int(np.floor(len(residuals) * 0.5)) - 1)
    if max_lags < 1:
        max_lags = 1

    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    axes[0, 0].plot(serie_modificada, marker='o', linestyle='-', label="Serie histórica")
    axes[0, 0].set_title(f"Serie histórica - {ing}")
    axes[0, 0].legend()
    axes[0, 1].plot(serie_modificada, marker='o', linestyle='-', label="Serie histórica")
    axes[0, 1].plot(forecast_index, prediccion, marker='o', linestyle='--', color='red', label="Pronóstico")
    axes[0, 1].set_title(f"Pronóstico - {ing}")
    axes[0, 1].legend()
    plot_acf(residuals, ax=axes[1, 0], lags=max_lags)
    axes[1, 0].set_title(f"ACF de residuos - {ing}")
    plot_pacf(residuals, ax=axes[1, 1], lags=max_lags)
    axes[1, 1].set_title(f"PACF de residuos - {ing}")
    plt.tight_layout()
    plt.show()

    # Crear DataFrame con la predicción para este ingrediente (fase 2)
    df_pred = pd.DataFrame({
         "Ingrediente": ing,
         "Fecha": forecast_index,
         "Consumo Predicho": prediccion.values
    })
    predicciones_list.append(df_pred)
    
    print(f"\n--- Pronóstico para {ing} ---")
    print(df_pred)
    
    # ---------------------------
    # Fase 3: Validación cruzada (walk-forward)
    # ---------------------------
    actual_cv, pred_cv = walk_forward_validation(serie_modificada, best_order, best_seasonal_order, forecast_horizon=1, initial_train_ratio=0.7)
    mae = mean_absolute_error(actual_cv, pred_cv)
    rmse = np.sqrt(mean_squared_error(actual_cv, pred_cv))
    mape = mean_absolute_percentage_error(actual_cv, pred_cv) * 100
    print(f"Validación cruzada para {ing}: MAE={mae:.2f}, RMSE={rmse:.2f}, MAPE={mape:.2f}%")
    
    # Graficar validación cruzada: valores reales vs pronosticados
    plt.figure(figsize=(10,4))
    plt.plot(actual_cv, marker='o', label="Valores reales")
    plt.plot(pred_cv, marker='o', linestyle='--', label="Walk-forward Forecast")
    plt.title(f"Validación Cruzada - {ing}")
    plt.xlabel("Punto de validación")
    plt.ylabel("Consumo")
    plt.legend()
    plt.show()

# Concatenar todas las predicciones en un solo DataFrame y crear tabla pivot
df_predicciones = pd.concat(predicciones_list, ignore_index=True)
df_predicciones["Año"] = df_predicciones["Fecha"].dt.year
df_predicciones["Mes"] = df_predicciones["Fecha"].dt.month
pivot_pred = df_predicciones.pivot(index=["Ingrediente", "Año"], columns="Mes", values="Consumo Predicho")
pivot_pred = pivot_pred.reindex(columns=range(1, 13), fill_value=0)
pivot_pred = pivot_pred.astype(float)
pivot_pred.rename(columns=meses_abreviados, inplace=True)
pivot_pred = pivot_pred.round(2)

ruta_pred = "prediccion_repeticion.xlsx"
with pd.ExcelWriter(ruta_pred, engine="openpyxl") as writer:
    pivot_pred.to_excel(writer, sheet_name="Predicciones", index=True)
    worksheet = writer.sheets["Predicciones"]
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[col_letter].width = max_length + 2
print(f"Archivo de predicciones generado: {ruta_pred}")
